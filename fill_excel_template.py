#!/usr/bin/env python3
"""
Script to fill Excel template with DuBox project structure information.
Scans backend and frontend, analyzes components, and fills the template.
"""

import os
import sys
import re
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import subprocess

# Try to import Excel libraries
try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not found. Attempting to install...")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "xlrd"])
        import openpyxl
        from openpyxl import load_workbook, Workbook
        HAS_OPENPYXL = True
    except:
        print("Error: Could not install openpyxl. Please install manually: pip install openpyxl xlrd")
        sys.exit(1)

# Configuration
EXCEL_TEMPLATE = r"C:\Users\DELL\Desktop\DUBOX Tracking Application.xls"
PROJECT_ROOT = r"C:\Users\DELL\source\repos\DuBox\DuBox-"
OUTPUT_FILE = os.path.join(PROJECT_ROOT, "Documentation", "Project_Filled.xlsx")

EXCLUDE_DIRS = {'node_modules', 'bin', 'obj', 'dist', '.angular', '.vscode', 
                'coverage', 'logs', 'wwwroot', 'packages', 'out-tsc'}

# Component analysis patterns
CONTROLLER_PATTERN = re.compile(r'class\s+(\w+Controller)\s*:')
COMPONENT_PATTERN = re.compile(r'@Component\s*\(\s*\{[^}]*selector:\s*[\'"](\w+)[\'"]')
SERVICE_PATTERN = re.compile(r'(?:class|export\s+class)\s+(\w+Service)\s*(?:extends|implements|:)')
ENTITY_PATTERN = re.compile(r'class\s+(\w+)\s*(?::|$)')
COMMAND_PATTERN = re.compile(r'(?:public\s+record\s+|class\s+)(\w+Command)\s*(?::|$)')
QUERY_PATTERN = re.compile(r'(?:public\s+record\s+|class\s+)(\w+Query)\s*(?::|$)')
HANDLER_PATTERN = re.compile(r'class\s+(\w+Handler)\s*:')
DTO_PATTERN = re.compile(r'(?:public\s+class|interface|export\s+interface)\s+(\w+Dto)\s*(?:extends|implements|:|{)')

def should_exclude_path(file_path: str) -> bool:
    """Check if file path should be excluded."""
    parts = Path(file_path).parts
    return any(part in EXCLUDE_DIRS for part in parts)

def scan_project_files() -> List[Dict]:
    """Scan project files and return list of file info."""
    project_files = []
    project_path = Path(PROJECT_ROOT)
    
    # Scan backend
    backend_paths = [
        project_path / "Dubox.Api",
        project_path / "Dubox.Application",
        project_path / "Dubox.Domain",
        project_path / "Dubox.Infrastructure"
    ]
    
    # Scan frontend
    frontend_path = project_path / "dubox-frontend" / "src" / "app"
    
    all_paths = backend_paths + [frontend_path]
    
    for base_path in all_paths:
        if not base_path.exists():
            continue
            
        for file_path in base_path.rglob("*"):
            if file_path.is_file() and not should_exclude_path(str(file_path)):
                rel_path = file_path.relative_to(PROJECT_ROOT)
                project_files.append({
                    'path': str(rel_path),
                    'full_path': str(file_path),
                    'extension': file_path.suffix.lower(),
                    'name': file_path.name
                })
    
    return project_files

def analyze_file(file_info: Dict) -> Optional[Dict]:
    """Analyze a file and extract component information."""
    try:
        with open(file_info['full_path'], 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
    except Exception as e:
        return None
    
    ext = file_info['extension']
    path = file_info['path']
    name = file_info['name']
    
    # Determine component type and purpose
    component_name = None
    purpose = ""
    what_it_does = ""
    dependencies = ""
    status = "Done"
    notes = ""
    
    # Backend Analysis
    if path.startswith("Dubox.Api/Controllers/"):
        match = CONTROLLER_PATTERN.search(content)
        if match:
            component_name = match.group(1)
            purpose = "API Controller"
            what_it_does = f"Handles HTTP requests for {component_name.replace('Controller', '')} operations"
            dependencies = "MediatR, Application Layer"
            # Extract endpoints
            endpoints = re.findall(r'\[(HttpGet|HttpPost|HttpPut|HttpDelete|HttpPatch)\]\s*\w+\s+(\w+)', content)
            if endpoints:
                notes = f"Endpoints: {', '.join([f'{m[0]}({m[1]})' for m in endpoints[:3]])}"
    
    elif path.startswith("Dubox.Application/Features/"):
        # Commands
        if "/Commands/" in path:
            match = COMMAND_PATTERN.search(content)
            if match:
                component_name = match.group(1)
                purpose = "CQRS Command"
                what_it_does = f"Defines command for {component_name.replace('Command', '')} operation"
                dependencies = "MediatR, Domain Layer"
        # Queries
        elif "/Queries/" in path:
            match = QUERY_PATTERN.search(content)
            if match:
                component_name = match.group(1)
                purpose = "CQRS Query"
                what_it_does = f"Defines query for retrieving {component_name.replace('Query', '')} data"
                dependencies = "MediatR, Domain Layer"
        # Handlers
        elif "Handler" in name:
            match = HANDLER_PATTERN.search(content)
            if match:
                component_name = match.group(1)
                purpose = "CQRS Handler"
                what_it_does = f"Handles {component_name.replace('Handler', '')} business logic"
                dependencies = "UnitOfWork, Repository, Domain Services"
        # DTOs
        elif path.startswith("Dubox.Application/DTOs/"):
            match = DTO_PATTERN.search(content)
            if match:
                component_name = match.group(1)
                purpose = "Data Transfer Object"
                what_it_does = f"Transfers {component_name.replace('Dto', '')} data between layers"
                dependencies = "Domain Entities"
    
    elif path.startswith("Dubox.Domain/Entities/"):
        match = ENTITY_PATTERN.search(content)
        if match:
            component_name = match.group(1)
            purpose = "Domain Entity"
            what_it_does = f"Represents {component_name} in the domain model"
            dependencies = "Domain Enums, Domain Interfaces"
    
    elif path.startswith("Dubox.Infrastructure/"):
        if "Repository" in name:
            component_name = name.replace(".cs", "")
            purpose = "Repository Implementation"
            what_it_does = f"Implements data access for {component_name}"
            dependencies = "Entity Framework, Domain Entities"
        elif "Service" in name:
            component_name = name.replace(".cs", "")
            purpose = "Infrastructure Service"
            what_it_does = f"Provides infrastructure service: {component_name}"
            dependencies = "External Libraries, Domain Interfaces"
    
    # Frontend Analysis
    elif path.startswith("dubox-frontend/src/app/"):
        if ext == ".ts" and "component.ts" in name:
            match = COMPONENT_PATTERN.search(content)
            if match:
                component_name = match.group(1)
                # Extract component class name
                class_match = re.search(r'export\s+class\s+(\w+Component)', content)
                if class_match:
                    component_name = class_match.group(1)
                
                purpose = "Angular Component"
                # Try to determine feature from path
                if "features/" in path:
                    feature = path.split("features/")[1].split("/")[0]
                    what_it_does = f"UI component for {feature} feature"
                else:
                    what_it_does = "Reusable UI component"
                dependencies = "Angular Core, Services, Models"
        
        elif ext == ".ts" and "service.ts" in name:
            match = SERVICE_PATTERN.search(content)
            if match:
                component_name = match.group(1)
                purpose = "Angular Service"
                what_it_does = f"Provides {component_name.replace('Service', '')} functionality to components"
                dependencies = "HttpClient, API Service"
        
        elif ext == ".ts" and "model.ts" in name:
            match = re.search(r'(?:export\s+)?(?:interface|class|type)\s+(\w+)', content)
            if match:
                component_name = match.group(1)
                purpose = "TypeScript Model/Interface"
                what_it_does = f"Defines {component_name} data structure"
                dependencies = "None"
        
        elif ext == ".ts" and "guard.ts" in name:
            match = re.search(r'export\s+(?:class|const)\s+(\w+Guard)', content)
            if match:
                component_name = match.group(1)
                purpose = "Angular Route Guard"
                what_it_does = f"Protects routes with {component_name.replace('Guard', '')} logic"
                dependencies = "Router, Auth Service"
        
        elif ext == ".ts" and "module.ts" in name:
            match = re.search(r'export\s+class\s+(\w+Module)', content)
            if match:
                component_name = match.group(1)
                purpose = "Angular Module"
                what_it_does = f"Organizes {component_name.replace('Module', '')} feature components"
                dependencies = "Angular Common, Feature Components"
    
    # If no specific pattern matched, try generic analysis
    if not component_name:
        # Try to extract class/interface name
        match = re.search(r'(?:export\s+)?(?:class|interface|type)\s+(\w+)', content)
        if match:
            component_name = match.group(1)
            purpose = "Code File"
            what_it_does = f"Contains {component_name} implementation"
            status = "Not Clear"
            notes = "Purpose unclear — needs clarification."
    
    if component_name:
        return {
            'component': component_name,
            'file_path': path,
            'purpose': purpose or "Code File",
            'what_it_does': what_it_does or f"Implements {component_name}",
            'dependencies': dependencies or "See code",
            'status': status,
            'notes': notes
        }
    
    return None

def read_excel_template(template_path: str) -> Tuple[Workbook, List[str]]:
    """Read Excel template and return workbook with column headers."""
    try:
        # Try to load as .xlsx first
        if template_path.endswith('.xls'):
            # Try to read as xlsx (some .xls files are actually xlsx)
            try:
                wb = load_workbook(template_path)
            except:
                # If that fails, we need xlrd for .xls files
                print("Warning: .xls file format detected. Converting to .xlsx format...")
                # For now, create a new workbook with standard columns
                wb = Workbook()
                ws = wb.active
                ws.title = "Project Components"
                headers = ['Component / Module', 'File / Path', 'Purpose', 
                          'What it does', 'Dependencies', 'Status', 'Notes']
                ws.append(headers)
                return wb, headers
        else:
            wb = load_workbook(template_path)
        
        # Get the first sheet
        ws = wb.active
        
        # Read headers from first row
        headers = []
        if ws.max_row > 0:
            for cell in ws[1]:
                headers.append(cell.value or "")
        
        # If no headers found, add default headers
        if not headers or all(not h for h in headers):
            headers = ['Component / Module', 'File / Path', 'Purpose', 
                      'What it does', 'Dependencies', 'Status', 'Notes']
            ws.append(headers)
        
        return wb, headers
    
    except Exception as e:
        print(f"Error reading template: {e}")
        # Create new workbook with standard structure
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Components"
        headers = ['Component / Module', 'File / Path', 'Purpose', 
                  'What it does', 'Dependencies', 'Status', 'Notes']
        ws.append(headers)
        return wb, headers

def fill_excel_template():
    """Main function to fill Excel template with project data."""
    print("Step 1: Reading Excel template...")
    if not os.path.exists(EXCEL_TEMPLATE):
        print(f"Error: Template not found at {EXCEL_TEMPLATE}")
        return False
    
    wb, headers = read_excel_template(EXCEL_TEMPLATE)
    ws = wb.active
    
    print("Step 2: Scanning project files...")
    project_files = scan_project_files()
    print(f"Found {len(project_files)} files to analyze")
    
    print("Step 3: Analyzing components...")
    components_data = []
    analyzed_count = 0
    
    for file_info in project_files:
        # Only analyze source code files
        if file_info['extension'] in ['.cs', '.ts', '.html', '.scss']:
            result = analyze_file(file_info)
            if result:
                components_data.append(result)
                analyzed_count += 1
                if analyzed_count % 50 == 0:
                    print(f"  Analyzed {analyzed_count} components...")
    
    print(f"Step 4: Filling Excel with {len(components_data)} components...")
    
    # Find column indices
    col_map = {}
    for i, header in enumerate(headers, start=1):
        header_lower = str(header).lower()
        if 'component' in header_lower or 'module' in header_lower:
            col_map['component'] = i
        elif 'file' in header_lower or 'path' in header_lower:
            col_map['file_path'] = i
        elif 'purpose' in header_lower:
            col_map['purpose'] = i
        elif 'what' in header_lower and 'does' in header_lower:
            col_map['what_it_does'] = i
        elif 'dependenc' in header_lower:
            col_map['dependencies'] = i
        elif 'status' in header_lower:
            col_map['status'] = i
        elif 'note' in header_lower:
            col_map['notes'] = i
    
    # Default column mapping if not found
    if not col_map:
        col_map = {
            'component': 1,
            'file_path': 2,
            'purpose': 3,
            'what_it_does': 4,
            'dependencies': 5,
            'status': 6,
            'notes': 7
        }
    
    # Find next empty row (after headers)
    next_row = ws.max_row + 1 if ws.max_row > 1 else 2
    
    # Fill data
    for comp in components_data:
        row = next_row
        if 'component' in col_map:
            ws.cell(row=row, column=col_map['component'], value=comp['component'])
        if 'file_path' in col_map:
            ws.cell(row=row, column=col_map['file_path'], value=comp['file_path'])
        if 'purpose' in col_map:
            ws.cell(row=row, column=col_map['purpose'], value=comp['purpose'])
        if 'what_it_does' in col_map:
            ws.cell(row=row, column=col_map['what_it_does'], value=comp['what_it_does'])
        if 'dependencies' in col_map:
            ws.cell(row=row, column=col_map['dependencies'], value=comp['dependencies'])
        if 'status' in col_map:
            ws.cell(row=row, column=col_map['status'], value=comp['status'])
        if 'notes' in col_map:
            ws.cell(row=row, column=col_map['notes'], value=comp['notes'])
        next_row += 1
    
    print("Step 5: Saving filled template...")
    # Create Documentation directory if it doesn't exist
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    
    # Save as .xlsx
    wb.save(OUTPUT_FILE)
    print(f"✓ Successfully saved to: {OUTPUT_FILE}")
    print(f"✓ Total components documented: {len(components_data)}")
    
    return True

if __name__ == "__main__":
    print("=" * 60)
    print("DuBox Project Structure Documentation Generator")
    print("=" * 60)
    success = fill_excel_template()
    if success:
        print("\n✓ Documentation complete!")
    else:
        print("\n✗ Documentation failed!")
        sys.exit(1)


